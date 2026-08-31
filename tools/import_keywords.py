#!/usr/bin/env python3
"""Build keywords.yaml from the keyword spreadsheet plus the manual layer.

    python tools/import_keywords.py                       # uses tools/keywords_source.xlsx
    python tools/import_keywords.py --xlsx ~/new_list.xlsx
    python tools/import_keywords.py --dry-run             # print, write nothing

What it does:
  * reads every sheet, treating row 1 as column headers (Records Management,
    Heritage, Mobility, Hospitality) and each column as a term list
  * merges the sheets, since they overlap heavily
  * splits "a / b" cells and "term (expansion)" cells into separate terms
  * applies the rewrite map (typos, ambiguous terms) from keywords_manual.yaml
  * routes terms to the right group by subject, not by which column they sat in
  * demotes terms that are too generic to carry a full group weight
  * folds in the French and German terms, negatives and CPV boosts
  * drops a term from every group but the highest-weighted one, so nothing is
    counted twice

Re-run it whenever the spreadsheet changes. It never touches config.yaml.
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "tools" / "keywords_source.xlsx"
DEFAULT_MANUAL = ROOT / "tools" / "keywords_manual.yaml"
DEFAULT_OUT = ROOT / "keywords.yaml"

# Spreadsheet column header -> group name. Headers are matched loosely.
COLUMN_GROUPS = {
    "records management": "records_management",
    "heritage": "heritage",
    "mobility": "mobility",
    "hospitality": "hospitality",
}

# Group ordering in the output, and the weight each carries. Groups not listed
# here are dropped.
GROUP_ORDER = ["records_management", "digitisation", "av_media", "heritage",
               "mobility", "fine_art", "hospitality", "support"]


def norm(text: str) -> str:
    """Comparison key: de-accented, lower-case, single-spaced."""
    text = unicodedata.normalize("NFKD", str(text))
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text).strip().lower()


def split_cell(value) -> list[str]:
    """One cell can hold several terms.

    'Digitization / Digitisation'                 -> both spellings
    'enterprise content management (ECM)'         -> phrase and acronym
    'Musems,Libraries Archives'                   -> handled by the rewrite map
    """
    if value is None:
        return []
    raw = str(value).replace("\u00a0", " ").strip()
    if not raw or raw.lower() in {"nan", "-", "–"}:
        return []

    parts: list[str] = []
    for chunk in re.split(r"\s*/\s*(?![^(]*\))", raw):
        chunk = chunk.strip(" -–—\t")
        if not chunk:
            continue
        m = re.match(r"^(.*?)\s*\(([^)]{2,})\)\s*$", chunk)
        if m:
            parts.extend([m.group(1).strip(), m.group(2).strip()])
        else:
            parts.append(chunk)
    return [p for p in parts if len(p) > 1]


def read_sheets(path: Path) -> dict[str, list[str]]:
    """Return {group_name: [terms]} from every sheet in the workbook."""
    wb = load_workbook(path, read_only=True, data_only=True)
    collected: dict[str, list[str]] = defaultdict(list)
    seen: set[tuple[str, str]] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [norm(c or "") for c in rows[0]]
        columns = {i: COLUMN_GROUPS[h] for i, h in enumerate(header) if h in COLUMN_GROUPS}
        if not columns:
            print(f"  skipped sheet '{sheet_name}': no recognised column headers",
                  file=sys.stderr)
            continue

        for row in rows[1:]:
            # Sheets can carry a second, prose taxonomy block below the term
            # table, separated by a blank row ("Core Records-Management
            # Process", "Content Types", ...). Those are section headings, not
            # search terms, so stop at the blank row.
            if not any(str(c).strip() for c in row if c is not None):
                break
            for idx, group in columns.items():
                if idx >= len(row):
                    continue
                cell = row[idx]
                # Bullet rows ("– Records Retrieval") belong to that same block.
                if cell is not None and str(cell).strip().startswith(("–", "—", "- ")):
                    continue
                for term in split_cell(cell):
                    key = (group, norm(term))
                    if key in seen:
                        continue
                    seen.add(key)
                    collected[group].append(term)
        print(f"  sheet '{sheet_name}': {len(columns)} columns read")
    return collected


def route(term: str, current_group: str, manual: dict) -> str:
    """Send a term to the group its subject belongs to."""
    key = norm(term)
    if key in {norm(d) for d in manual.get("demote", [])}:
        return "support"
    for rule in manual.get("routing", []):
        if any(needle in key for needle in (norm(m) for m in rule.get("match", []))):
            return rule["group"]
    return current_group


def build(xlsx: Path, manual_path: Path) -> tuple[dict, dict]:
    manual = yaml.safe_load(manual_path.read_text(encoding="utf-8")) or {}
    rewrite = {norm(k): v for k, v in (manual.get("rewrite") or {}).items()}

    print(f"Reading {xlsx.name}")
    sheet_terms = read_sheets(xlsx)

    groups: dict[str, list[str]] = defaultdict(list)
    stats = {"sheet_cells": 0, "rewritten": 0, "dropped": 0,
             "demoted": 0, "rerouted": 0, "deduped": 0, "subsumed": 0}

    for column_group, terms in sheet_terms.items():
        for term in terms:
            stats["sheet_cells"] += 1
            replacements = rewrite.get(norm(term))
            if replacements is not None:
                stats["rewritten"] += 1
                if not replacements:
                    stats["dropped"] += 1
                    continue
                candidates = replacements
            else:
                candidates = [term]

            for candidate in candidates:
                target = route(candidate, column_group, manual)
                if target == "support" and column_group != "support":
                    stats["demoted"] += 1
                elif target != column_group:
                    stats["rerouted"] += 1
                groups[target].append(candidate.strip())

    # Fold in the manual terms
    manual_groups = manual.get("groups") or {}
    weights = {name: float(cfg.get("weight", 1.0)) for name, cfg in manual_groups.items()}
    for name, cfg in manual_groups.items():
        groups[name].extend(cfg.get("terms") or [])

    for name in list(groups):
        weights.setdefault(name, 3.0)

    # Deduplicate: a term is kept only in its highest-weighted group, so
    # "scanning" in two groups does not score twice.
    best: dict[str, str] = {}
    for name in sorted(groups, key=lambda g: -weights[g]):
        for term in groups[name]:
            best.setdefault(norm(term), name)

    final: dict[str, list[str]] = defaultdict(list)
    emitted: set[str] = set()
    for name, terms in groups.items():
        for term in terms:
            key = norm(term)
            if key in emitted:
                stats["deduped"] += 1
                continue
            if best[key] != name:
                stats["deduped"] += 1
                continue
            emitted.add(key)
            final[name].append(key)

    # A wildcard term subsumes its own longer forms: keeping both "digitalisier*"
    # and "digitalisierung" would score the same word twice.
    for name, terms in final.items():
        stems = [t[:-1] for t in terms if t.endswith("*")]
        if not stems:
            continue
        kept = [t for t in terms
                if t.endswith("*") or not any(t.startswith(s) and t != s for s in stems)]
        stats["subsumed"] += len(terms) - len(kept)
        final[name] = kept

    out_groups = []
    for name in GROUP_ORDER:
        if final.get(name):
            out_groups.append({
                "name": name,
                "weight": weights.get(name, 3.0),
                "terms": sorted(set(final[name])),
            })
    for name in sorted(set(final) - set(GROUP_ORDER)):
        print(f"  note: group '{name}' is not in GROUP_ORDER, appending", file=sys.stderr)
        out_groups.append({"name": name, "weight": weights.get(name, 3.0),
                           "terms": sorted(set(final[name]))})

    negative = manual.get("negative") or {}
    document = {
        "groups": out_groups,
        "negative": {
            "weight": float(negative.get("weight", -7.0)),
            "terms": sorted({norm(t) for t in negative.get("terms", [])}),
        },
        "cpv_boost": manual.get("cpv_boost") or {},
    }
    return document, stats


HEADER = """# GENERATED FILE - do not edit by hand.
#
# Built from tools/keywords_source.xlsx + tools/keywords_manual.yaml by
#   python tools/import_keywords.py
#
# Add or change terms in tools/keywords_manual.yaml (French, negatives, CPV,
# routing, demotions) or in the spreadsheet (English), then re-run the importer.
# Terms are stored de-accented and lower-cased; "*" is a suffix wildcard.
"""


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--manual", default=str(DEFAULT_MANUAL))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    document, stats = build(Path(args.xlsx), Path(args.manual))

    total = sum(len(g["terms"]) for g in document["groups"])
    print(f"\n{stats['sheet_cells']} spreadsheet cells read")
    print(f"  {stats['rewritten']} rewritten, {stats['dropped']} dropped, "
          f"{stats['rerouted']} rerouted, {stats['demoted']} demoted to support")
    print(f"  {stats['deduped']} duplicates and {stats['subsumed']} wildcard-covered terms removed")
    print(f"\n{total} terms across {len(document['groups'])} groups:")
    for g in document["groups"]:
        print(f"  {g['name']:<20} weight {g['weight']:<5} {len(g['terms']):>4} terms")
    print(f"  {'negative':<20} weight {document['negative']['weight']:<5} "
          f"{len(document['negative']['terms']):>4} terms")

    body = yaml.safe_dump(document, allow_unicode=True, sort_keys=False, width=100)
    if args.dry_run:
        print("\n--- keywords.yaml (not written) ---\n")
        print(body)
        return 0

    Path(args.out).write_text(HEADER + body, encoding="utf-8")
    print(f"\nWrote {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
